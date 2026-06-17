#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Exporta o faturamento Omie localmente no formato usado pelo fluxo manual.

Saidas:
- Base exportada: 3 arquivos no layout dos pivots baixados do Omie.
- Base tratada: Produto_LS_*.xlsx e Vendedor_LS_*.xlsx.
- Final: copia de "Sheets/Base Faturamento.xlsx" com as abas principais preenchidas.
"""

from __future__ import annotations

import argparse
import copy
import html
import json
import os
import re
import sys
import time
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
ROOT_DIR = BASE_DIR.parent
OMIE_BASE_URL = "https://app.omie.com.br/api/v1"
HEADERS = {"Content-Type": "application/json"}

GREEN_FILL = PatternFill(start_color="00891A", end_color="00891A", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


@dataclass
class OmieAccount:
    name: str
    app_key: str
    app_secret: str


@dataclass
class AccountContext:
    account: OmieAccount
    client: "OmieClient"
    company_name: str
    company_cnpj: str
    vendors: dict[int, str]
    clients: dict[int, str]
    products: list[dict[str, Any]]
    coupons: list[dict[str, Any]]
    services: list[dict[str, Any]]
    finances: list[dict[str, Any]]


def load_env(path: Path) -> dict[str, str]:
    env: dict[str, str] = {}
    if not path.exists():
        return env

    for raw_line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        env[key] = value
    return env


def parse_accounts(env: dict[str, str]) -> list[OmieAccount]:
    accounts: list[OmieAccount] = []

    raw_credentials = env.get("OMIE_CREDENTIALS")
    if raw_credentials:
        try:
            credentials = json.loads(raw_credentials)
            if isinstance(credentials, dict):
                for name, item in credentials.items():
                    if not isinstance(item, dict):
                        continue
                    app_key = item.get("appKey") or item.get("app_key") or item.get("OMIE_APP_KEY")
                    app_secret = item.get("appSecret") or item.get("app_secret") or item.get("OMIE_APP_SECRET")
                    if app_key and app_secret:
                        accounts.append(
                            OmieAccount(
                                name=str(name),
                                app_key=str(app_key),
                                app_secret=str(app_secret),
                            )
                        )
            elif isinstance(credentials, list):
                for idx, item in enumerate(credentials, start=1):
                    if not isinstance(item, dict):
                        continue
                    app_key = item.get("appKey") or item.get("app_key") or item.get("OMIE_APP_KEY")
                    app_secret = item.get("appSecret") or item.get("app_secret") or item.get("OMIE_APP_SECRET")
                    if app_key and app_secret:
                        accounts.append(
                            OmieAccount(
                                name=str(item.get("name") or f"OMIE_CREDENTIALS_{idx}"),
                                app_key=str(app_key),
                                app_secret=str(app_secret),
                            )
                        )
        except json.JSONDecodeError as exc:
            raise ValueError(f"OMIE_CREDENTIALS invalido: {exc}") from exc

    if env.get("OMIE_APP_KEY") and env.get("OMIE_APP_SECRET"):
        accounts.append(
            OmieAccount(
                name=env.get("OMIE_ACCOUNT_NAME") or "OMIE_APP_KEY",
                app_key=env["OMIE_APP_KEY"],
                app_secret=env["OMIE_APP_SECRET"],
            )
        )

    raw_accounts = env.get("OMIE_ACCOUNTS_JSON")
    if raw_accounts:
        try:
            for idx, item in enumerate(json.loads(raw_accounts), start=1):
                app_key = item.get("app_key") or item.get("OMIE_APP_KEY")
                app_secret = item.get("app_secret") or item.get("OMIE_APP_SECRET")
                if app_key and app_secret:
                    accounts.append(
                        OmieAccount(
                            name=item.get("name") or f"OMIE_ACCOUNTS_JSON_{idx}",
                            app_key=app_key,
                            app_secret=app_secret,
                        )
                    )
        except json.JSONDecodeError as exc:
            raise ValueError(f"OMIE_ACCOUNTS_JSON invalido: {exc}") from exc

    for idx in range(1, 51):
        app_key = env.get(f"OMIE_APP_KEY_{idx}")
        app_secret = env.get(f"OMIE_APP_SECRET_{idx}")
        if app_key and app_secret:
            accounts.append(
                OmieAccount(
                    name=env.get(f"OMIE_ACCOUNT_NAME_{idx}") or f"OMIE_APP_KEY_{idx}",
                    app_key=app_key,
                    app_secret=app_secret,
                )
            )

    if not accounts:
        raise ValueError("Nenhuma credencial Omie encontrada no .env.")

    seen: set[tuple[str, str]] = set()
    unique_accounts: list[OmieAccount] = []
    for account in accounts:
        key = (account.app_key, account.app_secret)
        if key in seen:
            continue
        seen.add(key)
        unique_accounts.append(account)
    return unique_accounts


class OmieClient:
    def __init__(self, account: OmieAccount, timeout: int = 60):
        self.account = account
        self.timeout = timeout

    def call(self, endpoint: str, call: str, params: dict[str, Any] | None = None) -> dict[str, Any]:
        payload = {
            "call": call,
            "app_key": self.account.app_key,
            "app_secret": self.account.app_secret,
            "param": [params or {}],
        }
        for attempt in range(4):
            response = requests.post(
                f"{OMIE_BASE_URL}{endpoint}",
                headers=HEADERS,
                json=payload,
                timeout=self.timeout,
            )
            try:
                body = response.json()
            except ValueError:
                response.raise_for_status()
                return {}

            if response.status_code < 400 and "faultstring" not in body:
                return body

            message = body.get("faultstring") or body.get("message") or response.text
            if "Consumo redundante" in message and attempt < 3:
                match = re.search(r"Aguarde\s+(\d+)\s+segundos", message)
                wait_seconds = int(match.group(1)) + 1 if match else 10
                time.sleep(min(wait_seconds, 60))
                continue
            raise RuntimeError(f"{call} falhou: {message}")

        raise RuntimeError(f"{call} falhou sem resposta valida")

    def list_all(
        self,
        endpoint: str,
        call: str,
        list_key: str,
        params: dict[str, Any] | None = None,
        page_size: int = 100,
    ) -> list[dict[str, Any]]:
        params = dict(params or {})
        rows: list[dict[str, Any]] = []
        pagina = 1
        total_paginas = 1

        while pagina <= total_paginas:
            page_params = {
                "pagina": pagina,
                "registros_por_pagina": page_size,
                "apenas_importado_api": "N",
                **params,
            }
            try:
                response = self.call(endpoint, call, page_params)
            except RuntimeError as exc:
                if pagina == 1 and "Não existem registros" in str(exc):
                    return []
                raise

            rows.extend(response.get(list_key) or [])
            total_paginas = int(response.get("total_de_paginas") or response.get("nTotPaginas") or 1)
            pagina += 1

        return rows

    def list_all_numbered(
        self,
        endpoint: str,
        call: str,
        list_key: str,
        params: dict[str, Any] | None = None,
        page_size: int = 100,
    ) -> list[dict[str, Any]]:
        params = dict(params or {})
        rows: list[dict[str, Any]] = []
        pagina = 1
        total_paginas = 1

        while pagina <= total_paginas:
            page_params = {
                "nPagina": pagina,
                "nRegPorPagina": page_size,
                **params,
            }
            try:
                response = self.call(endpoint, call, page_params)
            except RuntimeError as exc:
                if pagina == 1 and "Não existem registros" in str(exc):
                    return []
                raise

            rows.extend(response.get(list_key) or [])
            total_paginas = int(response.get("nTotPaginas") or 1)
            pagina += 1

        return rows


def parse_date(value: str | None) -> date:
    if not value:
        return date.today() - timedelta(days=1)
    value = value.strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            pass
    raise ValueError("Data invalida. Use DD/MM/AAAA ou AAAA-MM-DD.")


def br_date(value: date) -> str:
    return value.strftime("%d/%m/%Y")


def file_date(value: date) -> str:
    return value.strftime("%d_%m_%Y")


def parse_omie_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def to_number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("R$", "").replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def clean_number(value: float) -> int | float:
    rounded = round(value, 2)
    if rounded.is_integer():
        return int(rounded)
    return rounded


def numeric_doc(value: Any) -> int | str:
    if value in (None, ""):
        return "N/D"
    text = str(value).strip()
    digits = re.sub(r"\D", "", text)
    if digits:
        return int(digits)
    return text


def document_root(value: Any) -> int | str:
    if value in (None, ""):
        return "N/D"
    text = str(value).strip()
    left = text.split("/", 1)[0]
    digits = re.sub(r"\D", "", left)
    if digits:
        return int(digits)
    digits = re.sub(r"\D", "", text)
    if len(digits) > 3 and digits.endswith("001"):
        return int(digits[:-3])
    if digits:
        return int(digits)
    return text


def document_display(value: Any) -> str:
    if value in (None, ""):
        return "N/D"
    return str(value).strip()


def padded_doc(value: Any, size: int) -> str:
    if value in (None, ""):
        return "N/D"
    digits = re.sub(r"\D", "", str(value))
    if not digits:
        return str(value)
    return digits.zfill(size)


def normalize_key(value: Any) -> str:
    return str(value or "").strip().casefold()


def clean_description(value: Any) -> str:
    text = html.unescape(str(value or "").strip())
    if not text:
        return ""
    letters = [char for char in text if char.isalpha()]
    if letters and all(not char.islower() for char in letters):
        text = text.title()
        replacements = {
            "Usb": "USB",
            "Hdmi": "HDMI",
            "P2": "P2",
            "Iphone": "iPhone",
            "Ipad": "iPad",
            "Imac": "iMac",
            "Macbook": "Macbook",
            "Apple Watch": "Apple Watch",
            "Nmve": "NMVE",
            "Ssd": "SSD",
        }
        for old, new in replacements.items():
            text = text.replace(old, new)
    return text


def get_nested(mapping: dict[str, Any], *keys: str) -> Any:
    current: Any = mapping
    for key in keys:
        if not isinstance(current, dict):
            return None
        current = current.get(key)
    return current


def collect_reference_data(client: OmieClient) -> tuple[str, str, dict[int, str]]:
    company_name = client.account.name
    company_cnpj = ""
    empresas = client.list_all("/geral/empresas/", "ListarEmpresas", "empresas_cadastro", page_size=100)
    if empresas:
        company_name = empresas[0].get("nome_fantasia") or empresas[0].get("razao_social") or company_name
        company_cnpj = empresas[0].get("cnpj") or ""

    vendors_rows = client.list_all("/geral/vendedores/", "ListarVendedores", "cadastro", page_size=100)
    vendors = {
        int(row["codigo"]): row.get("nome", "")
        for row in vendors_rows
        if row.get("codigo") is not None
    }
    return company_name, company_cnpj, vendors


def get_client_name(ctx: AccountContext, codigo_cliente: Any) -> str:
    if codigo_cliente in (None, ""):
        return ""
    try:
        codigo = int(codigo_cliente)
    except (TypeError, ValueError):
        return str(codigo_cliente)
    if codigo in ctx.clients:
        return ctx.clients[codigo]

    try:
        response = ctx.client.call("/geral/clientes/", "ConsultarCliente", {"codigo_cliente_omie": codigo})
        name = response.get("razao_social") or response.get("nome_fantasia") or str(codigo)
    except Exception:
        name = str(codigo)
    ctx.clients[codigo] = name
    return name


def get_vendor_name(ctx: AccountContext, codigo_vendedor: Any) -> str:
    try:
        codigo = int(codigo_vendedor)
    except (TypeError, ValueError):
        return str(codigo_vendedor or "")
    return ctx.vendors.get(codigo, str(codigo))


def fetch_context(account: OmieAccount, target_date: date) -> AccountContext:
    client = OmieClient(account)
    company_name, company_cnpj, vendors = collect_reference_data(client)
    data = br_date(target_date)

    products = client.list_all(
        "/produtos/pedido/",
        "ListarPedidos",
        "pedido_venda_produto",
        {
            "data_faturamento_de": data,
            "data_faturamento_ate": data,
            "status_pedido": "FATURADO",
        },
        page_size=50,
    )

    services = client.list_all(
        "/servicos/os/",
        "ListarOS",
        "osCadastro",
        {
            "filtrar_por_data_faturamento_de": data,
            "filtrar_por_data_faturamento_ate": data,
            "filtrar_por_status": "F",
        },
        page_size=50,
    )

    coupons = client.list_all_numbered(
        "/produtos/cupomfiscalconsultar/",
        "CuponsFiscais",
        "cupons",
        {
            "dDtEmissaoDe": data,
            "dDtEmissaoAte": data,
        },
        page_size=50,
    )

    finances = client.list_all(
        "/financas/contareceber/",
        "ListarContasReceber",
        "conta_receber_cadastro",
        {
            "ordenar_por": "DATA_EMISSAO",
            "filtrar_por_emissao_de": data,
            "filtrar_por_emissao_ate": data,
        },
        page_size=100,
    )

    return AccountContext(
        account=account,
        client=client,
        company_name=company_name,
        company_cnpj=company_cnpj,
        vendors=vendors,
        clients={},
        products=products,
        coupons=coupons,
        services=services,
        finances=finances,
    )


def product_nfe_number(ctx: AccountContext, pedido: dict[str, Any]) -> tuple[str, str]:
    codigo_pedido = get_nested(pedido, "cabecalho", "codigo_pedido")
    if not codigo_pedido:
        return "N/D", ""

    try:
        status = ctx.client.call("/produtos/pedido/", "StatusPedido", {"codigo_pedido": codigo_pedido})
    except Exception:
        return "N/D", ""

    nfes = status.get("ListaNfe") or []
    if not nfes:
        return "N/D", ""
    nfe = nfes[0]
    return nfe.get("numero_nfe") or "N/D", nfe.get("data_emissao") or ""


def collapse_product_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[Any, ...], dict[str, Any]] = {}
    for row in rows:
        key = (
            row.get("date"),
            normalize_key(row.get("company")),
            row.get("invoice"),
            normalize_key(row.get("description")),
            normalize_key(row.get("vendor")),
        )
        if key not in grouped:
            grouped[key] = dict(row)
            grouped[key]["amount"] = to_number(row.get("amount"))
            continue
        grouped[key]["amount"] = to_number(grouped[key].get("amount")) + to_number(row.get("amount"))

    collapsed = []
    for row in grouped.values():
        row["amount"] = clean_number(to_number(row.get("amount")))
        collapsed.append(row)
    return collapsed


def build_rows(contexts: list[AccountContext]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    product_rows: list[dict[str, Any]] = []
    service_rows: list[dict[str, Any]] = []
    vendor_rows: list[dict[str, Any]] = []

    for ctx in contexts:
        canceled_coupon_docs: set[int | str] = set()
        finance_by_os: dict[int, dict[str, Any]] = {}
        for conta in ctx.finances:
            if conta.get("nCodOS") is not None:
                try:
                    finance_by_os[int(conta["nCodOS"])] = conta
                except (TypeError, ValueError):
                    pass

        for pedido in ctx.products:
            nf_number, nf_date = product_nfe_number(ctx, pedido)
            date_value = parse_omie_date(nf_date) or parse_omie_date(get_nested(pedido, "infoCadastro", "dFat"))
            vendor = get_vendor_name(ctx, get_nested(pedido, "informacoes_adicionais", "codVend"))
            company = ctx.company_name
            for item in pedido.get("det") or []:
                produto = item.get("produto") or {}
                codigo = str(produto.get("codigo") or "").strip()
                descricao = clean_description(produto.get("descricao"))
                if codigo and descricao:
                    descricao_completa = f"{codigo} - {descricao}"
                else:
                    descricao_completa = descricao or codigo
                product_rows.append(
                    {
                        "date": date_value,
                        "company": company,
                        "invoice": numeric_doc(nf_number),
                        "invoice_raw": padded_doc(nf_number, 8),
                        "description": descricao_completa,
                        "vendor": vendor,
                        "amount": clean_number(to_number(produto.get("valor_total"))),
                    }
                )

        for cupom in ctx.coupons:
            header = cupom.get("cabecalhoCupom") or {}
            info = header.get("info") or {}
            if info.get("cCupomCancelado") == "S" or info.get("cCupomDevolvido") == "S":
                canceled_coupon_docs.add(document_root(header.get("nNumCupom")))
                continue
            date_value = parse_omie_date(header.get("dDtEmissaoCupom"))
            vendor = get_vendor_name(ctx, header.get("idVendedor"))
            invoice = numeric_doc(header.get("nNumCupom"))
            invoice_raw = padded_doc(header.get("nNumCupom"), 8)
            for item in cupom.get("itensCupom") or []:
                if item.get("cItemCancelado") == "S" or item.get("cItemDevolvido") == "S":
                    continue
                codigo = str(item.get("cCodigo") or item.get("emiProduto") or "").strip()
                descricao = clean_description(item.get("xProd"))
                if codigo and descricao:
                    descricao_completa = f"{codigo} - {descricao}"
                else:
                    descricao_completa = descricao or codigo
                product_rows.append(
                    {
                        "date": date_value,
                        "company": ctx.company_name,
                        "invoice": invoice,
                        "invoice_raw": invoice_raw,
                        "description": descricao_completa,
                        "vendor": vendor,
                        "amount": clean_number(to_number(item.get("vItem"))),
                    }
                )

        os_by_id: dict[int, dict[str, Any]] = {}
        for ordem in ctx.services:
            ncod_os = get_nested(ordem, "Cabecalho", "nCodOS")
            if ncod_os is not None:
                try:
                    os_by_id[int(ncod_os)] = ordem
                except (TypeError, ValueError):
                    pass
            conta = finance_by_os.get(int(ncod_os)) if ncod_os is not None and str(ncod_os).isdigit() else None
            date_value = parse_omie_date(get_nested(ordem, "InfoCadastro", "dDtFat"))
            vendor = get_vendor_name(ctx, get_nested(ordem, "Cabecalho", "nCodVend"))
            client_name = get_client_name(
                ctx,
                (conta or {}).get("codigo_cliente_fornecedor") or get_nested(ordem, "Cabecalho", "nCodCli"),
            )
            nfs_number = (conta or {}).get("numero_documento_fiscal") or "N/D"
            for servico in sorted(ordem.get("ServicosPrestados") or [], key=lambda row: str(row.get("cDescServ") or "")):
                qty = to_number(servico.get("nQtde") or 1)
                unit = to_number(servico.get("nValUnit"))
                discount = to_number(servico.get("nValorDesconto"))
                additions = to_number(servico.get("nValorAcrescimos"))
                gross = clean_number(qty * unit)
                total = clean_number(qty * unit + additions - discount)
                service_rows.append(
                    {
                        "date": date_value,
                        "company": ctx.company_name,
                        "company_cnpj": ctx.company_cnpj,
                        "nfs": numeric_doc(nfs_number),
                        "nfs_raw": padded_doc(nfs_number, 13),
                        "receipt": "N/D",
                        "vendor": vendor,
                        "client": client_name,
                        "description": clean_description(servico.get("cDescServ")),
                        "amount": gross,
                        "discount": clean_number(discount),
                        "total": total,
                        "liquid": total,
                        "category_code": servico.get("cCodCategItem") or "",
                        "os_code": get_nested(ordem, "Cabecalho", "nCodOS"),
                    }
                )

        for conta in ctx.finances:
            if document_root(conta.get("numero_documento_fiscal") or conta.get("numero_documento")) in canceled_coupon_docs:
                continue

            if conta.get("nCodOS") is not None:
                try:
                    ordem = os_by_id.get(int(conta["nCodOS"]))
                except (TypeError, ValueError):
                    ordem = None

                if ordem:
                    grouped: dict[str, float] = {}
                    for servico in ordem.get("ServicosPrestados") or []:
                        code = servico.get("cCodCategItem") or "N/D"
                        qty = to_number(servico.get("nQtde") or 1)
                        unit = to_number(servico.get("nValUnit"))
                        discount = to_number(servico.get("nValorDesconto"))
                        additions = to_number(servico.get("nValorAcrescimos"))
                        grouped[code] = grouped.get(code, 0.0) + qty * unit + additions - discount

                    for _, amount in sorted(grouped.items()):
                        vendor_rows.append(
                            {
                                "date": parse_omie_date(conta.get("data_emissao")),
                                "document": document_display(conta.get("numero_documento_fiscal")),
                                "document_number": numeric_doc(conta.get("numero_documento_fiscal")),
                                "vendor": get_vendor_name(ctx, conta.get("codigo_vendedor")),
                                "client": get_client_name(ctx, conta.get("codigo_cliente_fornecedor")),
                                "amount": clean_number(amount),
                                "type": "1. Contas a Receber",
                                "company": ctx.company_name,
                            }
                        )
                    continue

            vendor_rows.append(
                {
                    "date": parse_omie_date(conta.get("data_emissao")),
                    "document": document_display(conta.get("numero_documento_fiscal") or conta.get("numero_documento")),
                    "document_number": numeric_doc(conta.get("numero_documento_fiscal") or conta.get("numero_documento")),
                    "vendor": get_vendor_name(ctx, conta.get("codigo_vendedor")),
                    "client": get_client_name(ctx, conta.get("codigo_cliente_fornecedor")),
                    "amount": clean_number(to_number(conta.get("valor_documento"))),
                    "type": "1. Contas a Receber",
                    "company": ctx.company_name,
                }
            )

    product_rows = collapse_product_rows(product_rows)
    product_rows.sort(key=lambda row: (row["date"] or date.min, normalize_key(row["company"]), row["invoice"] if isinstance(row["invoice"], int) else 0, normalize_key(row["description"])))
    service_rows.sort(key=lambda row: (row["date"] or date.min, normalize_key(row["company"]), normalize_key(row["description"])))
    vendor_rows.sort(key=lambda row: (row["date"] or date.min, normalize_key(row["company"]), row.get("document_number") if isinstance(row.get("document_number"), int) else 0, normalize_key(row["vendor"]), row["amount"]))
    return product_rows, service_rows, vendor_rows


def apply_header_style(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = GREEN_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def autosize(ws, max_width: int = 45) -> None:
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        width = 12
        for cell in column_cells:
            if cell.value is None:
                continue
            width = max(width, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[letter].width = width


def save_table(rows: list[dict[str, Any]], headers: list[str], keys: list[str], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados"
    ws.append(headers)
    for row in rows:
        ws.append([row.get(key) for key in keys])
    apply_header_style(ws)
    ws.freeze_panes = "A2"
    autosize(ws)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_base_exportada(
    product_rows: list[dict[str, Any]],
    service_rows: list[dict[str, Any]],
    vendor_rows: list[dict[str, Any]],
    output_dir: Path,
) -> dict[str, str]:
    output_dir.mkdir(parents=True, exist_ok=True)
    files: dict[str, str] = {}

    wb = Workbook()
    ws = wb.active
    ws.title = "Faturamento ProdutoEmpresa ..."
    ws.merge_cells("A1:F1")
    ws["A1"] = "Faturamento ProdutoEmpresa Faturamento por Período"
    ws["B2"] = "Filtros"
    ws["A3"] = "Situação"
    ws["B3"] = "Autorizado"
    ws.append([])
    ws.append([
        "Data de Emissão (completa)",
        "Minha Empresa (Nome Fantasia)",
        "Nota Fiscal",
        "Descrição do Produto (completa)",
        "Vendedor",
        "Total da Nota Fiscal",
    ])
    for row in product_rows:
        ws.append([
            br_date(row["date"]) if row["date"] else "",
            row["company"],
            row["invoice_raw"],
            row["description"],
            row["vendor"],
            row["amount"],
        ])
    autosize(ws)
    path = output_dir / "pivot.xlsx"
    wb.save(path)
    files["produto_exportado"] = str(path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Faturamento por Período For..."
    ws.merge_cells("A1:J1")
    ws["A1"] = "Faturamento por Período Forma de Pagamento Vendedor Faturamento por Período"
    ws["B2"] = "Filtros"
    ws["A3"] = "Situação"
    ws["B3"] = "NFS-e Autorizada"
    if service_rows:
        ws["A4"] = "Minha Empresa (Nome Fantasia)"
        ws["B4"] = service_rows[0]["company"]
    ws.append([])
    ws.append([
        "Minha Empresa (CNPJ)",
        "Situação",
        "Data de Emissão (completa)",
        "Número da NFS-e",
        "Número do Recibo",
        "Vendedor",
        "Cliente (Razão Social)",
        "Descrição do Serviço (resumida)",
        "Valor do Serviço",
        "Valor do Desconto",
        "Valor Total do Serviço",
        "Valor Líquido",
    ])
    for row in service_rows:
        ws.append([
            row["company_cnpj"],
            "NFS-e Autorizada",
            br_date(row["date"]) if row["date"] else "",
            row["nfs_raw"],
            row["receipt"],
            row["vendor"],
            row["client"],
            row["description"],
            row["amount"],
            row["discount"],
            row["total"],
            row["liquid"],
        ])
    if service_rows:
        ws.append(["Total geral", "", "", "", "", "", "", "", sum(to_number(row["amount"]) for row in service_rows), sum(to_number(row["discount"]) for row in service_rows), sum(to_number(row["total"]) for row in service_rows), sum(to_number(row["liquid"]) for row in service_rows)])
    autosize(ws)
    path = output_dir / "pivot (1).xlsx"
    wb.save(path)
    files["servico_exportado"] = str(path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Faturamento por Vendedor Em..."
    ws.merge_cells("A1:G1")
    ws["A1"] = "Faturamento por Vendedor Empresa Contas por Vendedor"
    ws.append([
        "Data de Emissão (completa)",
        "NF/CF",
        "Vendedor",
        "Cliente ou Fornecedor (Nome Fantasia)",
        "Valor da Conta",
        "Tipo",
        "Nome do Meu Aplicativo",
    ])
    ws.append(["", "", "", "", clean_number(sum(to_number(row["amount"]) for row in vendor_rows)), "", ""])
    for row in vendor_rows:
        ws.append([
            br_date(row["date"]) if row["date"] else "",
            row["document"],
            row["vendor"],
            row["client"],
            row["amount"],
            row["type"],
            row["company"],
        ])
    autosize(ws)
    path = output_dir / "pivot (2).xlsx"
    wb.save(path)
    files["vendedor_exportado"] = str(path)
    return files


def write_treated(
    product_rows: list[dict[str, Any]],
    service_rows: list[dict[str, Any]],
    vendor_rows: list[dict[str, Any]],
    output_dir: Path,
    run_date: date,
) -> dict[str, str]:
    output_dir.mkdir(parents=True, exist_ok=True)
    suffix = file_date(run_date)
    files: dict[str, str] = {}

    produto_rows = [
        {
            "Data de Emissão (completa)": row["date"],
            "Minha Empresa (Nome Fantasia)": row["company"],
            "Nota Fiscal": row["invoice"],
            "Descrição do Produto (completa)": row["description"],
            "Vendedor": row["vendor"],
            "Total da Nota Fiscal": row["amount"],
        }
        for row in product_rows
    ] + [
        {
            "Data de Emissão (completa)": row["date"],
            "Minha Empresa (Nome Fantasia)": row["company"],
            "Nota Fiscal": "N/D",
            "Descrição do Produto (completa)": row["description"],
            "Vendedor": row["vendor"],
            "Total da Nota Fiscal": row["total"],
        }
        for row in service_rows
    ]

    produto_path = output_dir / f"Produto_LS_{suffix}.xlsx"
    save_table(
        produto_rows,
        [
            "Data de Emissão (completa)",
            "Minha Empresa (Nome Fantasia)",
            "Nota Fiscal",
            "Descrição do Produto (completa)",
            "Vendedor",
            "Total da Nota Fiscal",
        ],
        [
            "Data de Emissão (completa)",
            "Minha Empresa (Nome Fantasia)",
            "Nota Fiscal",
            "Descrição do Produto (completa)",
            "Vendedor",
            "Total da Nota Fiscal",
        ],
        produto_path,
    )
    files["produto_tratado"] = str(produto_path)

    vendedor_rows = [
        {
            "Data de Emissão (completa)": row["date"],
            "NF/CF": row["document"],
            "Vendedor": row["vendor"],
            "Cliente ou Fornecedor (Nome Fantasia)": row["client"],
            "Valor da Conta": row["amount"],
            "Tipo": row["type"],
            "Nome do Meu Aplicativo": row["company"],
        }
        for row in vendor_rows
    ]
    vendedor_path = output_dir / f"Vendedor_LS_{suffix}.xlsx"
    save_table(
        vendedor_rows,
        [
            "Data de Emissão (completa)",
            "NF/CF",
            "Vendedor",
            "Cliente ou Fornecedor (Nome Fantasia)",
            "Valor da Conta",
            "Tipo",
            "Nome do Meu Aplicativo",
        ],
        [
            "Data de Emissão (completa)",
            "NF/CF",
            "Vendedor",
            "Cliente ou Fornecedor (Nome Fantasia)",
            "Valor da Conta",
            "Tipo",
            "Nome do Meu Aplicativo",
        ],
        vendedor_path,
    )
    files["vendedor_tratado"] = str(vendedor_path)
    return files


def copy_row_style(ws, source_row: int, target_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy.copy(source._style)
        target.number_format = source.number_format
        target.alignment = copy.copy(source.alignment)
        target.border = copy.copy(source.border)
        target.fill = copy.copy(source.fill)
        target.font = copy.copy(source.font)


def sheet_date_matches(value: Any, target_date: date) -> bool:
    parsed = parse_omie_date(value)
    return parsed == target_date


def delete_existing_rows(ws, target_date: date, company_col: int, companies: set[str]) -> int:
    deleted = 0
    for row_idx in range(ws.max_row, 1, -1):
        if not sheet_date_matches(ws.cell(row_idx, 1).value, target_date):
            continue
        company = normalize_key(ws.cell(row_idx, company_col).value)
        if company in companies:
            ws.delete_rows(row_idx, 1)
            deleted += 1
    return deleted


def read_lookup_sheets(template: Path) -> dict[str, Any]:
    wb = load_workbook(template, data_only=True)

    product_terms: list[tuple[str, str]] = []
    ws = wb["class_produto"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            product_terms.append((normalize_key(row[0]), str(row[1])))

    device_terms: list[tuple[str, str]] = []
    ws = wb["class_device"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            device_terms.append((normalize_key(row[0]), str(row[1])))

    type_by_category: dict[str, str] = {}
    ws = wb["class_Tipo"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            type_by_category[normalize_key(row[0])] = str(row[1])

    access_by_vendor: dict[str, str] = {}
    ws = wb["_Colaborador"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            access_by_vendor[normalize_key(row[0])] = str(row[1])

    sector_by_vendor: dict[str, str] = {}
    ws = wb["Colaboradores"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[3]:
            sector_by_vendor[normalize_key(row[0])] = str(row[3])

    return {
        "product_terms": product_terms,
        "device_terms": device_terms,
        "type_by_category": type_by_category,
        "access_by_vendor": access_by_vendor,
        "sector_by_vendor": sector_by_vendor,
    }


def classify(text: str, terms: list[tuple[str, str]]) -> str:
    normalized = normalize_key(text)
    for needle, category in terms:
        if needle and needle in normalized:
            return category
    return ""


def write_final_workbook(
    product_rows: list[dict[str, Any]],
    service_rows: list[dict[str, Any]],
    vendor_rows: list[dict[str, Any]],
    template: Path,
    output_dir: Path,
    target_date: date,
    run_date: date,
    replace_date: bool,
) -> dict[str, str]:
    if not template.exists():
        return {}

    lookups = read_lookup_sheets(template)
    wb = load_workbook(template, data_only=False)
    output_dir.mkdir(parents=True, exist_ok=True)
    companies = {normalize_key(row["company"]) for row in product_rows + service_rows + vendor_rows}
    deleted = {"produtos": 0, "vendedor": 0}

    product_final_rows: list[dict[str, Any]] = []
    for row in product_rows:
        category = classify(row["description"], lookups["product_terms"])
        product_type = lookups["type_by_category"].get(normalize_key(category), "")
        device = classify(row["description"], lookups["device_terms"])
        product_final_rows.append({**row, "category": category, "type": product_type, "device": device, "invoice_for_final": row["invoice"]})
    for row in service_rows:
        category = classify(row["description"], lookups["product_terms"])
        product_type = lookups["type_by_category"].get(normalize_key(category), "Serviço")
        device = classify(row["description"], lookups["device_terms"])
        product_final_rows.append({**row, "category": category, "type": product_type, "device": device, "invoice_for_final": "N/D", "amount": row["total"]})

    assistance_by_doc: dict[tuple[str, Any], str] = {}
    for row in product_final_rows:
        key = (normalize_key(row["company"]), row["invoice_for_final"])
        if row["type"] == "Serviço":
            assistance_by_doc[key] = "Sim"
        elif key not in assistance_by_doc:
            assistance_by_doc[key] = "Não"

    ws = wb["Produtos e Servicos"]
    if replace_date:
        deleted["produtos"] = delete_existing_rows(ws, target_date, 2, companies)
    style_row = max(ws.max_row, 2)
    for row in product_final_rows:
        target_row = ws.max_row + 1
        copy_row_style(ws, style_row, target_row, 11)
        invoice = row["invoice_for_final"]
        ws.cell(target_row, 1).value = datetime.combine(row["date"], datetime.min.time()) if row.get("date") else ""
        ws.cell(target_row, 2).value = row["company"]
        ws.cell(target_row, 3).value = invoice
        ws.cell(target_row, 4).value = row["description"]
        ws.cell(target_row, 5).value = row["vendor"]
        ws.cell(target_row, 6).value = row["amount"]
        ws.cell(target_row, 7).value = assistance_by_doc.get((normalize_key(row["company"]), invoice), "Não")
        ws.cell(target_row, 8).value = row["type"]
        ws.cell(target_row, 9).value = row["category"]
        ws.cell(target_row, 10).value = row["device"]
        ws.cell(target_row, 11).value = lookups["access_by_vendor"].get(normalize_key(row["vendor"]), "#N/A")

    ws = wb["Vendedor"]
    if replace_date:
        deleted["vendedor"] = delete_existing_rows(ws, target_date, 7, companies)
    style_row = max(ws.max_row, 2)
    for row in vendor_rows:
        target_row = ws.max_row + 1
        copy_row_style(ws, style_row, target_row, 12)
        document = row["document"]
        document_number = row.get("document_number", numeric_doc(document))
        sector = lookups["sector_by_vendor"].get(normalize_key(row["vendor"]), row["company"])
        category = "comercial" if "Comercial - Sede" in sector else "unidades"
        ws.cell(target_row, 1).value = datetime.combine(row["date"], datetime.min.time()) if row.get("date") else ""
        ws.cell(target_row, 2).value = document
        ws.cell(target_row, 3).value = row["vendor"]
        ws.cell(target_row, 4).value = row["client"]
        ws.cell(target_row, 5).value = row["amount"]
        ws.cell(target_row, 6).value = row["type"]
        ws.cell(target_row, 7).value = row["company"]
        ws.cell(target_row, 8).value = document_number
        ws.cell(target_row, 9).value = assistance_by_doc.get((normalize_key(row["company"]), document_number), "Não")
        ws.cell(target_row, 10).value = sector
        ws.cell(target_row, 11).value = category
        ws.cell(target_row, 12).value = lookups["access_by_vendor"].get(normalize_key(row["vendor"]), "#N/A")

    final_path = output_dir / f"Base_Faturamento_OMIE_{file_date(run_date)}.xlsx"
    wb.save(final_path)
    return {"base_faturamento_final": str(final_path), "deleted_rows": deleted}


def write_manifest(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Exporta Omie local no formato das bases LS.")
    parser.add_argument("--date", dest="target_date", help="Data faturada/emissao a exportar. Padrao: ontem.")
    parser.add_argument("--run-date", dest="run_date", help="Data usada no sufixo dos arquivos. Padrao: hoje.")
    parser.add_argument("--env", default=str(ROOT_DIR / ".env"), help="Caminho do .env.")
    parser.add_argument("--output", default=str(BASE_DIR / "output"), help="Pasta base de saida.")
    parser.add_argument("--template", default=str(BASE_DIR / "Sheets" / "Base Faturamento.xlsx"), help="Template da planilha final.")
    parser.add_argument("--no-final", action="store_true", help="Nao gera a copia final da planilha.")
    parser.add_argument("--append-only", action="store_true", help="Nao remove linhas existentes da mesma data/empresa no template final.")
    args = parser.parse_args(argv)

    target_date = parse_date(args.target_date)
    run_date = parse_date(args.run_date) if args.run_date else date.today()
    env = {**os.environ, **load_env(Path(args.env))}
    accounts = parse_accounts(env)
    day_dir = Path(args.output) / target_date.strftime("%Y-%m-%d")

    print(f"[omie-local] Data alvo: {br_date(target_date)}")
    print(f"[omie-local] Contas Omie configuradas: {len(accounts)}")

    contexts: list[AccountContext] = []
    for idx, account in enumerate(accounts, start=1):
        print(f"[omie-local] [{idx}/{len(accounts)}] Coletando conta {account.name}...")
        ctx = fetch_context(account, target_date)
        contexts.append(ctx)
        print(
            "[omie-local] "
            f"{ctx.company_name}: pedidos={len(ctx.products)}, cupons={len(ctx.coupons)}, "
            f"os={len(ctx.services)}, contas={len(ctx.finances)}"
        )

    product_rows, service_rows, vendor_rows = build_rows(contexts)
    print(
        "[omie-local] Linhas tratadas: "
        f"produtos={len(product_rows)}, servicos={len(service_rows)}, vendedor={len(vendor_rows)}"
    )

    files: dict[str, Any] = {}
    files.update(write_base_exportada(product_rows, service_rows, vendor_rows, day_dir / "Base exportada"))
    files.update(write_treated(product_rows, service_rows, vendor_rows, day_dir / "Base tratada", run_date))
    if not args.no_final:
        files.update(
            write_final_workbook(
                product_rows,
                service_rows,
                vendor_rows,
                Path(args.template),
                day_dir / "Sheets",
                target_date,
                run_date,
                replace_date=not args.append_only,
            )
        )

    manifest = {
        "target_date": br_date(target_date),
        "run_date": br_date(run_date),
        "accounts": [
            {
                "name": ctx.account.name,
                "company_name": ctx.company_name,
                "products": len(ctx.products),
                "coupons": len(ctx.coupons),
                "services": len(ctx.services),
                "finances": len(ctx.finances),
            }
            for ctx in contexts
        ],
        "rows": {
            "product_rows": len(product_rows),
            "service_rows": len(service_rows),
            "vendor_rows": len(vendor_rows),
        },
        "files": files,
    }
    manifest_path = day_dir / "manifest.json"
    write_manifest(manifest_path, manifest)
    print(f"[omie-local] Manifest: {manifest_path}")
    for label, file_path in files.items():
        if label == "deleted_rows":
            print(f"[omie-local] Linhas substituidas no template: {file_path}")
        else:
            print(f"[omie-local] {label}: {file_path}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"[omie-local] ERRO: {exc}", file=sys.stderr)
        raise SystemExit(1)
